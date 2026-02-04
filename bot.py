import asyncio
import logging
import re
import sqlite3
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime, date
from io import BytesIO
from typing import Optional, Dict, Any, List

from aiogram import Bot, Dispatcher, types, F
from aiogram.client.default import DefaultBotProperties
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton, BufferedInputFile

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# =========================
# SETTINGS (НЕ env, как ты хотел)
# =========================
BOT_TOKEN = "8524573448:AAFHW_KFr8_M2iJ5tp5rZRSPkj_GVjM-V34"
DATABASE_FILE = "attendance.db"

ADMIN_IDS: List[int] = [7233585816]
SECRET_CODE = "учитель2026"

MAX_STUDENTS = 32

# -100... id приватной группы (бот должен быть в группе и иметь право писать)
REPORT_CHAT_ID: Optional[int] = -1003756818645

LOG_LEVEL = "INFO"

# Для процентов под ИТОГО
SHIFT_TOTAL_STUDENTS = {
    False: 1067,  # 1 смена
    True: 256,    # 2 смена
}


# =========================
# LOGGING
# =========================
logging.basicConfig(
    level=getattr(logging, LOG_LEVEL, logging.INFO),
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    handlers=[logging.FileHandler("bot.log", encoding="utf-8"), logging.StreamHandler()],
)
logger = logging.getLogger("attendance_bot")


# =========================
# BOT / DISPATCHER
# =========================
bot = Bot(token=BOT_TOKEN, default=DefaultBotProperties(parse_mode="HTML"))
dp = Dispatcher(storage=MemoryStorage())

# анти-лаг: один юзер = одна операция callback за раз
USER_LOCKS: Dict[int, asyncio.Lock] = defaultdict(asyncio.Lock)

# анти-даблклик: уже обработанные callback key (ограниченный кеш)
SEEN_CALLBACKS: set[str] = set()


# =========================
# FSM
# =========================
class AdminStates(StatesGroup):
    waiting_for_subject = State()
    waiting_for_class = State()
    waiting_for_broadcast = State()
    waiting_for_search = State()


# =========================
# DB
# =========================
def get_connection() -> sqlite3.Connection:
    conn = sqlite3.connect(DATABASE_FILE, timeout=30)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    conn.execute("PRAGMA journal_mode = WAL")
    conn.execute("PRAGMA synchronous = NORMAL")
    conn.execute("PRAGMA busy_timeout = 5000")
    return conn


def ensure_column(conn: sqlite3.Connection, table: str, column: str, col_def: str):
    cur = conn.cursor()
    cur.execute(f"PRAGMA table_info({table})")
    cols = {row[1] for row in cur.fetchall()}
    if column not in cols:
        cur.execute(f"ALTER TABLE {table} ADD COLUMN {column} {col_def}")
        conn.commit()


def init_db():
    conn = get_connection()
    cur = conn.cursor()

    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS teachers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            telegram_id INTEGER UNIQUE NOT NULL,
            full_name TEXT NOT NULL,
            subject TEXT,
            class_name TEXT,
            is_approved INTEGER DEFAULT 0,
            registered_at TEXT DEFAULT (datetime('now'))
        )
        """
    )

    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS daily_reports (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            teacher_id INTEGER NOT NULL,
            report_date TEXT NOT NULL,

            sick INTEGER DEFAULT 0,
            sanatorium INTEGER DEFAULT 0,
            parent_statement INTEGER DEFAULT 0,
            competition INTEGER DEFAULT 0,
            absent_without_reason INTEGER DEFAULT 0,

            is_submitted INTEGER DEFAULT 0,
            submitted_at TEXT,

            FOREIGN KEY (teacher_id) REFERENCES teachers (id) ON DELETE CASCADE,
            UNIQUE(teacher_id, report_date)
        )
        """
    )

    cur.execute("CREATE INDEX IF NOT EXISTS idx_teachers_telegram ON teachers(telegram_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_teachers_approved ON teachers(is_approved)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_reports_date ON daily_reports(report_date)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_reports_teacher ON daily_reports(teacher_id)")

    ensure_column(conn, "teachers", "subject", "TEXT")
    ensure_column(conn, "teachers", "class_name", "TEXT")

    conn.commit()
    conn.close()
    logger.info("✅ База данных инициализирована")


# =========================
# HELPERS
# =========================
def today_iso() -> str:
    return date.today().isoformat()


def safe_int(x: Any, default: int = 0) -> int:
    try:
        return int(x)
    except Exception:
        return default


def is_admin(user_id: int) -> bool:
    return user_id in ADMIN_IDS


async def safe_answer(cb: types.CallbackQuery, text: Optional[str] = None, show_alert: bool = False):
    try:
        await cb.answer(text=text, show_alert=show_alert)
    except Exception:
        pass


async def safe_edit_text(msg: types.Message, text: str, reply_markup: Optional[InlineKeyboardMarkup] = None):
    try:
        await msg.edit_text(text, reply_markup=reply_markup)
    except Exception:
        try:
            await msg.answer(text, reply_markup=reply_markup)
        except Exception:
            pass


def grade_from_class_name(class_name: Optional[str]) -> Optional[int]:
    if not class_name:
        return None
    s = str(class_name).strip()
    m = re.match(r"^\s*(\d+)", s)
    if not m:
        return None
    try:
        return int(m.group(1))
    except Exception:
        return None


def is_second_shift_class(class_name: Optional[str]) -> bool:
    # 2 смена = 3* и 6*
    g = grade_from_class_name(class_name)
    return g in {3, 6}


def cb_seen_key(cb: types.CallbackQuery) -> str:
    # ключ на случай "двойного" апдейта
    return f"{cb.from_user.id}|{cb.id}|{cb.data}"


# =========================
# DB QUERIES
# =========================
def get_teacher_by_telegram(telegram_id: int) -> Optional[Dict[str, Any]]:
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT * FROM teachers WHERE telegram_id = ?", (telegram_id,))
    row = cur.fetchone()
    conn.close()
    return dict(row) if row else None


def create_teacher_if_missing(telegram_id: int, full_name: str) -> Dict[str, Any]:
    t = get_teacher_by_telegram(telegram_id)
    if t:
        return t
    conn = get_connection()
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO teachers(telegram_id, full_name, is_approved) VALUES (?, ?, 0)",
        (telegram_id, full_name),
    )
    conn.commit()
    conn.close()
    return get_teacher_by_telegram(telegram_id) or {"telegram_id": telegram_id, "full_name": full_name}


def get_teacher_status(telegram_id: int) -> str:
    if is_admin(telegram_id):
        return "admin"
    t = get_teacher_by_telegram(telegram_id)
    if not t:
        return "guest"
    return "approved_teacher" if safe_int(t.get("is_approved")) == 1 else "pending_teacher"


def approve_teacher_db(teacher_id: int):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("UPDATE teachers SET is_approved = 1 WHERE id = ?", (teacher_id,))
    conn.commit()
    conn.close()


def decline_teacher_db(teacher_id: int):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM teachers WHERE id = ?", (teacher_id,))
    conn.commit()
    conn.close()


def update_teacher_field(teacher_id: int, field: str, value: str):
    if field not in {"subject", "class_name", "full_name"}:
        return
    conn = get_connection()
    cur = conn.cursor()
    cur.execute(f"UPDATE teachers SET {field} = ? WHERE id = ?", (value, teacher_id))
    conn.commit()
    conn.close()


def get_pending_teachers() -> List[Dict[str, Any]]:
    conn = get_connection()
    cur = conn.cursor()
    cur.execute(
        "SELECT id, telegram_id, full_name, subject, class_name FROM teachers WHERE is_approved = 0 ORDER BY registered_at DESC"
    )
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return rows


def get_approved_teachers() -> List[Dict[str, Any]]:
    conn = get_connection()
    cur = conn.cursor()
    cur.execute(
        "SELECT id, telegram_id, full_name, subject, class_name FROM teachers WHERE is_approved = 1 ORDER BY class_name, full_name"
    )
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return rows


def search_teachers(q: str) -> List[Dict[str, Any]]:
    q = (q or "").strip()
    if not q:
        return []
    q_low = q.lower()
    like = f"%{q_low}%"

    conn = get_connection()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT id, telegram_id, full_name, subject, class_name, is_approved
        FROM teachers
        WHERE lower(full_name) LIKE ?
           OR lower(class_name) LIKE ?
           OR CAST(telegram_id AS TEXT) LIKE ?
           OR CAST(id AS TEXT) LIKE ?
        ORDER BY is_approved DESC, class_name, full_name
        """,
        (like, like, f"%{q}%", f"%{q}%"),
    )
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return rows


def get_shift_teachers(second_shift: bool) -> List[Dict[str, Any]]:
    rows = get_approved_teachers()
    return [t for t in rows if is_second_shift_class(t.get("class_name")) == second_shift]


def create_or_get_report(teacher_id: int, report_date_iso: str) -> Dict[str, Any]:
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT * FROM daily_reports WHERE teacher_id = ? AND report_date = ?", (teacher_id, report_date_iso))
    row = cur.fetchone()
    if row:
        conn.close()
        return dict(row)

    cur.execute("INSERT INTO daily_reports(teacher_id, report_date) VALUES(?, ?)", (teacher_id, report_date_iso))
    conn.commit()

    cur.execute("SELECT * FROM daily_reports WHERE teacher_id = ? AND report_date = ?", (teacher_id, report_date_iso))
    row2 = cur.fetchone()
    conn.close()
    return dict(row2) if row2 else {}


def get_report_by_teacher_id(teacher_id: int, report_date_iso: str) -> Optional[Dict[str, Any]]:
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT * FROM daily_reports WHERE teacher_id = ? AND report_date = ?", (teacher_id, report_date_iso))
    row = cur.fetchone()
    conn.close()
    return dict(row) if row else None


def set_report_value(report_id: int, field: str, value: int) -> bool:
    if field not in {"sick", "sanatorium", "parent_statement", "competition", "absent_without_reason"}:
        return False
    value = max(0, min(int(MAX_STUDENTS), safe_int(value)))
    conn = get_connection()
    cur = conn.cursor()
    cur.execute(f"UPDATE daily_reports SET {field} = ? WHERE id = ?", (value, report_id))
    conn.commit()
    ok = cur.rowcount > 0
    conn.close()
    return ok


def submit_report(report_id: int) -> bool:
    conn = get_connection()
    cur = conn.cursor()
    cur.execute(
        "UPDATE daily_reports SET is_submitted = 1, submitted_at = datetime('now') WHERE id = ? AND is_submitted = 0",
        (report_id,),
    )
    conn.commit()
    ok = cur.rowcount > 0
    conn.close()
    return ok


def create_today_reports_for_all_approved():
    d = today_iso()
    for t in get_approved_teachers():
        create_or_get_report(int(t["id"]), d)


# =========================
# KEYBOARDS
# =========================
def admin_main_kb() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(text="👥 Ожидают", callback_data="admin_pending"),
                InlineKeyboardButton(text="✅ Подтвержденные", callback_data="admin_approved"),
            ],
            [
                InlineKeyboardButton(text="🔍 Поиск", callback_data="admin_search"),
                InlineKeyboardButton(text="📢 Рассылка", callback_data="admin_broadcast"),
            ],
            [InlineKeyboardButton(text="🔄 Обновить", callback_data="admin_refresh")],
        ]
    )


def back_kb(cb: str = "admin_back") -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text="🔙 Назад", callback_data=cb)]])


def report_kb(report: Dict[str, Any]) -> InlineKeyboardMarkup:
    rid = int(report["id"])
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="🤒 Болеют", callback_data=f"edit|sick|{rid}")],
            [InlineKeyboardButton(text="🏥 На оздоровлении (санаторий)", callback_data=f"edit|sanatorium|{rid}")],
            [InlineKeyboardButton(text="📄 По заявлению родителей", callback_data=f"edit|parent_statement|{rid}")],
            [InlineKeyboardButton(text="🏆 На соревнованиях", callback_data=f"edit|competition|{rid}")],
            [InlineKeyboardButton(text="❌ Без уважительной причины", callback_data=f"edit|absent_without_reason|{rid}")],
            [
                InlineKeyboardButton(text="📤 Отправить отчет", callback_data=f"submit|{rid}"),
                InlineKeyboardButton(text="🔄 Обновить", callback_data=f"refresh|{rid}"),
            ],
        ]
    )


def number_kb(field: str, current: int, report_id: int) -> InlineKeyboardMarkup:
    max_n = int(MAX_STUDENTS)
    numbers = list(range(0, max_n + 1))

    rows: List[List[InlineKeyboardButton]] = []
    row: List[InlineKeyboardButton] = []

    for n in numbers:
        txt = f"✅ {n}" if n == current else str(n)
        row.append(InlineKeyboardButton(text=txt, callback_data=f"set|{field}|{n}|{report_id}"))
        if len(row) == 6:
            rows.append(row)
            row = []
    if row:
        rows.append(row)

    rows.append(
        [
            InlineKeyboardButton(text=f"🎯 Сейчас: {current}", callback_data="noop"),
            InlineKeyboardButton(text="+1", callback_data=f"delta|{field}|1|{report_id}"),
            InlineKeyboardButton(text="+5", callback_data=f"delta|{field}|5|{report_id}"),
        ]
    )
    rows.append(
        [
            InlineKeyboardButton(text="Сброс 0", callback_data=f"set|{field}|0|{report_id}"),
            InlineKeyboardButton(text="🔙 Назад", callback_data=f"back|{report_id}"),
        ]
    )
    return InlineKeyboardMarkup(inline_keyboard=rows)


def teacher_fill_kb() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text="📊 Заполнить отчет", callback_data="manual_report_start")]])


# =========================
# REPORT VIEW
# =========================
def report_text(rep: Dict[str, Any]) -> str:
    return (
        "📊 Заполните отчет по отсутствующим:\n\n"
        f"🤒 Болеют: <b>{safe_int(rep.get('sick'))}</b>\n"
        f"🏥 На оздоровлении (санаторий): <b>{safe_int(rep.get('sanatorium'))}</b>\n"
        f"📄 По заявлению родителей: <b>{safe_int(rep.get('parent_statement'))}</b>\n"
        f"🏆 На соревнованиях: <b>{safe_int(rep.get('competition'))}</b>\n"
        f"❌ Без уважительной причины: <b>{safe_int(rep.get('absent_without_reason'))}</b>"
    )


async def show_report(chat_id: int, telegram_id: int):
    t = get_teacher_by_telegram(telegram_id)
    if not t or safe_int(t.get("is_approved")) != 1:
        await bot.send_message(chat_id, "❌ Вы не подтвержденный учитель.")
        return

    create_today_reports_for_all_approved()
    rep = create_or_get_report(int(t["id"]), today_iso())

    if safe_int(rep.get("is_submitted")) == 1:
        await bot.send_message(chat_id, "✅ Вы уже отправили отчет на сегодня.")
        return

    await bot.send_message(chat_id, report_text(rep), reply_markup=report_kb(rep))


async def notify_teacher_fill(telegram_id: int, report_date_iso: str, label: str):
    try:
        await bot.send_message(
            telegram_id,
            f"🔔 <b>{label}</b>\n\nПожалуйста, заполните отчет за <b>{report_date_iso}</b>.",
            reply_markup=teacher_fill_kb(),
        )
    except Exception as e:
        logger.error(f"notify_teacher_fill error to {telegram_id}: {e}")


# =========================
# EXCEL SUMMARY (без учителя/предмета + проценты)
# =========================
def build_shift_excel(report_date_iso: str, second_shift: bool, teachers: List[Dict[str, Any]]) -> bytes:
    title = "2 смена (3* и 6*)" if second_shift else "1 смена (кроме 3* и 6*)"
    denom = int(SHIFT_TOTAL_STUDENTS.get(second_shift, 1)) or 1

    wb = Workbook()
    ws = wb.active
    ws.title = "Отчеты"

    header_fill = PatternFill("solid", fgColor="E7EEF7")
    total_fill = PatternFill("solid", fgColor="FFF2CC")
    percent_fill = PatternFill("solid", fgColor="E2F0D9")

    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    thin = Side(style="thin", color="A0A0A0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = f"Итоговые отчеты за {report_date_iso} — {title}"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:G1")
    ws.row_dimensions[1].height = 24

    headers = [
        "Класс",
        "Статус",
        "Болеют",
        "Санаторий",
        "По заявлению родителей",
        "На соревнованиях",
        "Без уважительной причины",
    ]
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=2, column=col)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[2].height = 34

    total_sick = total_san = total_parent = total_comp = total_no = 0
    submitted_cnt = 0
    pending_cnt = 0

    row = 3
    for t in teachers:
        rep = get_report_by_teacher_id(int(t["id"]), report_date_iso)
        cls = t.get("class_name") or "—"

        if rep and safe_int(rep.get("is_submitted")) == 1:
            status = "Сдан"
            submitted_cnt += 1
            sick = safe_int(rep.get("sick"))
            san = safe_int(rep.get("sanatorium"))
            par = safe_int(rep.get("parent_statement"))
            comp = safe_int(rep.get("competition"))
            no = safe_int(rep.get("absent_without_reason"))

            total_sick += sick
            total_san += san
            total_parent += par
            total_comp += comp
            total_no += no
        else:
            status = "Не сдан"
            pending_cnt += 1
            sick = san = par = comp = no = 0

        ws.append([cls, status, sick, san, par, comp, no])

        for c in range(1, 8):
            cell = ws.cell(row=row, column=c)
            cell.border = border
            cell.alignment = left if c == 1 else center
        row += 1

    # ИТОГО
    ws.append(
        [
            "ИТОГО",
            f"Сдано: {submitted_cnt} / Не сдано: {pending_cnt}",
            total_sick,
            total_san,
            total_parent,
            total_comp,
            total_no,
        ]
    )
    total_row = row
    for c in range(1, 8):
        cell = ws.cell(row=total_row, column=c)
        cell.font = bold
        cell.fill = total_fill
        cell.border = border
        cell.alignment = center if c >= 2 else left

    # ПРОЦЕНТЫ (каждый пункт / 1067 или / 256)
    ws.append(
        [
            "ПРОЦЕНТ",
            f"от {denom}",
            total_sick / denom,
            total_san / denom,
            total_parent / denom,
            total_comp / denom,
            total_no / denom,
        ]
    )
    percent_row = total_row + 1
    for c in range(1, 8):
        cell = ws.cell(row=percent_row, column=c)
        cell.font = bold
        cell.fill = percent_fill
        cell.border = border
        cell.alignment = center if c >= 2 else left

    for c in range(3, 8):
        ws.cell(row=percent_row, column=c).number_format = "0.00%"

    # ширины
    col_widths = [10, 22, 10, 12, 22, 18, 24]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A3"

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


async def send_shift_excel_to_chat(report_date_iso: str, second_shift: bool, teachers: List[Dict[str, Any]]):
    file_bytes = build_shift_excel(report_date_iso, second_shift, teachers)
    shift_label = "2смена" if second_shift else "1смена"
    filename = f"Отчеты_{report_date_iso}_{shift_label}.xlsx"
    doc = BufferedInputFile(file_bytes, filename=filename)

    if REPORT_CHAT_ID is not None:
        try:
            await bot.send_document(REPORT_CHAT_ID, doc, caption=f"📎 Итог {report_date_iso} ({shift_label})")
        except Exception as e:
            logger.error(f"send_shift_excel_to_chat error: {e}")
        return

    for admin_id in ADMIN_IDS:
        try:
            await bot.send_document(admin_id, doc, caption=f"📎 Итог {report_date_iso} ({shift_label})")
        except Exception as e:
            logger.error(f"send_excel_to_admin error: {e}")


# =========================
# SCHEDULE (как ты хотел)
# =========================
@dataclass(frozen=True)
class ScheduleEvent:
    hhmm: str
    second_shift: bool
    label: str
    is_summary: bool


SCHEDULE: List[ScheduleEvent] = [
    ScheduleEvent("07:50", False, "Напоминание 1/4", False),
    ScheduleEvent("08:10", False, "Напоминание 2/4", False),
    ScheduleEvent("08:20", False, "Напоминание 3/4", False),
    ScheduleEvent("08:30", False, "Напоминание 4/4", False),
    ScheduleEvent("08:40", False, "Итог", True),

    ScheduleEvent("13:10", True, "Напоминание 1/3", False),
    ScheduleEvent("13:20", True, "Напоминание 2/3", False),
    ScheduleEvent("13:30", True, "Напоминание 3/3", False),
    ScheduleEvent("13:40", True, "Итог", True),
]


_EXECUTED: set[str] = set()


def minute_key(d_iso: str, hhmm: str, second_shift: bool) -> str:
    return f"{d_iso}|{hhmm}|{'2' if second_shift else '1'}"


async def schedule_loop():
    while True:
        try:
            now = datetime.now()
            d_iso = now.date().isoformat()
            hhmm = now.strftime("%H:%M")

            for ev in SCHEDULE:
                if ev.hhmm != hhmm:
                    continue

                key = minute_key(d_iso, ev.hhmm, ev.second_shift)
                if key in _EXECUTED:
                    continue
                _EXECUTED.add(key)

                create_today_reports_for_all_approved()
                teachers = get_shift_teachers(ev.second_shift)

                if ev.is_summary:
                    await send_shift_excel_to_chat(d_iso, ev.second_shift, teachers)
                else:
                    for t in teachers:
                        rep = get_report_by_teacher_id(int(t["id"]), d_iso)
                        if rep and safe_int(rep.get("is_submitted")) == 1:
                            continue
                        await notify_teacher_fill(int(t["telegram_id"]), d_iso, ev.label)
                        await asyncio.sleep(0.05)

            if len(_EXECUTED) > 10000:
                _EXECUTED.clear()

            await asyncio.sleep(2)
        except Exception as e:
            logger.error(f"schedule_loop error: {e}")
            await asyncio.sleep(5)


# =========================
# COMMANDS
# =========================
@dp.message(Command("start"))
async def cmd_start(message: types.Message):
    status = get_teacher_status(message.from_user.id)

    if status == "approved_teacher":
        await message.answer(
            "👋 Добро пожаловать!\n\n"
            "📋 Команды:\n"
            "/report — заполнить отчет\n"
            "/status — статус отчета\n"
            "/profile — мой профиль\n"
            "/help — помощь"
        )
    elif status == "pending_teacher":
        await message.answer("⏳ Ваша заявка отправлена и ожидает подтверждения администратора.")
    elif status == "admin":
        await message.answer("👑 Администратор. Используйте /admin.")
    else:
        await message.answer("🤖 Введите секретный код для регистрации учителя.")


@dp.message(Command("help"))
async def cmd_help(message: types.Message):
    status = get_teacher_status(message.from_user.id)
    if status == "admin":
        await message.answer(
            "📚 Админ: /admin\n"
            "Учитель: /report /status /profile\n"
            "Регистрация: отправить секретный код.\n\n"
            "Тест: /chatid /testtable"
        )
    else:
        await message.answer(
            "📚 Команды учителя: /report /status /profile\n"
            "Регистрация: отправьте секретный код."
        )


@dp.message(Command("report"))
async def cmd_report(message: types.Message):
    await show_report(message.chat.id, message.from_user.id)


@dp.message(Command("status"))
async def cmd_status(message: types.Message):
    t = get_teacher_by_telegram(message.from_user.id)
    if not t or safe_int(t.get("is_approved")) != 1:
        await message.answer("⛔ Доступно только подтвержденным учителям.")
        return

    rep = get_report_by_teacher_id(int(t["id"]), today_iso())
    if not rep:
        await message.answer("📅 Отчет на сегодня еще не создан.")
        return

    await message.answer("✅ Отчет отправлен." if safe_int(rep.get("is_submitted")) == 1 else "⏳ Отчет не отправлен. Нажмите /report")


@dp.message(Command("profile"))
async def cmd_profile(message: types.Message):
    t = get_teacher_by_telegram(message.from_user.id)
    if not t:
        await message.answer("Профиль не найден.")
        return
    await message.answer(
        "👤 Профиль:\n"
        f"ФИО: <b>{t.get('full_name')}</b>\n"
        f"Предмет: <b>{t.get('subject') or '—'}</b>\n"
        f"Класс: <b>{t.get('class_name') or '—'}</b>\n"
        f"Статус: <b>{'✅ подтвержден' if safe_int(t.get('is_approved')) == 1 else '⏳ ожидает'}</b>"
    )


@dp.message(Command("chatid"))
async def cmd_chatid(message: types.Message):
    await message.answer(f"chat_id = <code>{message.chat.id}</code>")


@dp.message(Command("testtable"))
async def cmd_testtable(message: types.Message):
    if not is_admin(message.from_user.id):
        await message.answer("Нет доступа")
        return
    create_today_reports_for_all_approved()
    d = today_iso()
    await send_shift_excel_to_chat(d, False, get_shift_teachers(False))
    await send_shift_excel_to_chat(d, True, get_shift_teachers(True))
    await message.answer("✅ Excel отправлен(ы).")


# =========================
# ADMIN PANEL
# =========================
@dp.message(Command("admin"))
async def cmd_admin(message: types.Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        await message.answer("⛔ Нет доступа")
        return
    await admin_render(message, state)


@dp.callback_query(F.data == "admin_back")
async def cb_admin_back(cb: types.CallbackQuery, state: FSMContext):
    if not is_admin(cb.from_user.id):
        await safe_answer(cb, "⛔ Нет доступа", show_alert=True)
        return
    await admin_render(cb.message, state)
    await safe_answer(cb)


async def admin_render(msg: types.Message, state: FSMContext):
    create_today_reports_for_all_approved()
    today = today_iso()

    conn = get_connection()
    cur = conn.cursor()

    cur.execute("SELECT COUNT(*) AS c FROM teachers WHERE is_approved = 1")
    approved_count = cur.fetchone()["c"]

    cur.execute("SELECT COUNT(*) AS c FROM teachers WHERE is_approved = 0")
    pending_count = cur.fetchone()["c"]

    cur.execute(
        "SELECT COUNT(*) AS total, SUM(CASE WHEN is_submitted=1 THEN 1 ELSE 0 END) AS submitted FROM daily_reports WHERE report_date = ?",
        (today,),
    )
    st = cur.fetchone()
    conn.close()

    text = (
        "👑 Админ-панель\n\n"
        f"📅 Сегодня: <b>{today}</b>\n"
        f"✅ Подтвержденных: <b>{approved_count}</b>\n"
        f"⏳ Ожидают: <b>{pending_count}</b>\n"
        f"📤 Отчетов сдано: <b>{st['submitted'] or 0}</b> / <b>{st['total'] or 0}</b>\n\n"
        "Выберите действие:"
    )
    await msg.answer(text, reply_markup=admin_main_kb())


@dp.callback_query(F.data == "admin_refresh")
async def admin_refresh(cb: types.CallbackQuery, state: FSMContext):
    if not is_admin(cb.from_user.id):
        await safe_answer(cb, "⛔ Нет доступа", show_alert=True)
        return
    await safe_answer(cb, "✅")
    await cb.message.answer("🔄 Обновлено. /admin")


@dp.callback_query(F.data == "admin_pending")
async def admin_pending(cb: types.CallbackQuery):
    if not is_admin(cb.from_user.id):
        await safe_answer(cb, "⛔ Нет доступа", show_alert=True)
        return

    rows = get_pending_teachers()
    if not rows:
        await safe_edit_text(cb.message, "👥 Ожидающих нет.", back_kb())
        await safe_answer(cb)
        return

    lines = ["👥 Ожидают подтверждения:\n"]
    kb_rows: List[List[InlineKeyboardButton]] = []
    for t in rows[:25]:
        cls = t.get("class_name") or "—"
        subj = t.get("subject") or "—"
        lines.append(f"• <b>{t['full_name']}</b> (id: <code>{t['id']}</code>, {cls}, {subj})")
        kb_rows.append([InlineKeyboardButton(text=f"✅ Подтвердить {t['id']}", callback_data=f"appr|{t['id']}")])
        kb_rows.append([InlineKeyboardButton(text=f"❌ Отклонить {t['id']}", callback_data=f"decl|{t['id']}")])
        kb_rows.append([InlineKeyboardButton(text=f"✏️ Класс {t['id']}", callback_data=f"setclass|{t['id']}")])
        kb_rows.append([InlineKeyboardButton(text=f"✏️ Предмет {t['id']}", callback_data=f"setsubj|{t['id']}")])

    kb_rows.append([InlineKeyboardButton(text="🔙 Назад", callback_data="admin_back")])
    await safe_edit_text(cb.message, "\n".join(lines), InlineKeyboardMarkup(inline_keyboard=kb_rows))
    await safe_answer(cb)


@dp.callback_query(F.data == "admin_approved")
async def admin_approved(cb: types.CallbackQuery):
    if not is_admin(cb.from_user.id):
        await safe_answer(cb, "⛔ Нет доступа", show_alert=True)
        return

    rows = get_approved_teachers()
    if not rows:
        await safe_edit_text(cb.message, "✅ Подтвержденных нет.", back_kb())
        await safe_answer(cb)
        return

    lines = ["✅ Подтвержденные:\n"]
    kb_rows: List[List[InlineKeyboardButton]] = []
    for t in rows[:35]:
        cls = t.get("class_name") or "—"
        subj = t.get("subject") or "—"
        lines.append(f"• <b>{cls}</b> — {t['full_name']} (id: <code>{t['id']}</code>, {subj})")
        kb_rows.append([InlineKeyboardButton(text=f"✏️ Изменить класс {t['id']}", callback_data=f"setclass|{t['id']}")])
        kb_rows.append([InlineKeyboardButton(text=f"✏️ Изменить предмет {t['id']}", callback_data=f"setsubj|{t['id']}")])

    kb_rows.append([InlineKeyboardButton(text="🔙 Назад", callback_data="admin_back")])
    await safe_edit_text(cb.message, "\n".join(lines), InlineKeyboardMarkup(inline_keyboard=kb_rows))
    await safe_answer(cb)


@dp.callback_query(F.data == "admin_search")
async def admin_search(cb: types.CallbackQuery, state: FSMContext):
    if not is_admin(cb.from_user.id):
        await safe_answer(cb, "⛔ Нет доступа", show_alert=True)
        return
    await state.set_state(AdminStates.waiting_for_search)
    await safe_answer(cb)
    await cb.message.answer("Введите запрос для поиска (ФИО / класс / teacher_id / telegram_id):")


@dp.message(AdminStates.waiting_for_search)
async def st_wait_search(message: types.Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return

    q = (message.text or "").strip()
    if not q:
        await message.answer("Введите непустой запрос.")
        return

    rows = search_teachers(q)
    await state.clear()

    if not rows:
        await message.answer("Ничего не найдено. /admin")
        return

    lines = [f"🔍 Результаты ({len(rows)}):\n"]
    kb_rows: List[List[InlineKeyboardButton]] = []

    for t in rows[:25]:
        cls = t.get("class_name") or "—"
        subj = t.get("subject") or "—"
        st = "✅" if safe_int(t.get("is_approved")) == 1 else "⏳"
        lines.append(f"{st} <b>{cls}</b> — {t['full_name']} (id: <code>{t['id']}</code>, {subj})")

        if safe_int(t.get("is_approved")) == 0:
            kb_rows.append([InlineKeyboardButton(text=f"✅ Подтвердить {t['id']}", callback_data=f"appr|{t['id']}")])
            kb_rows.append([InlineKeyboardButton(text=f"❌ Отклонить {t['id']}", callback_data=f"decl|{t['id']}")])

        kb_rows.append([InlineKeyboardButton(text=f"✏️ Класс {t['id']}", callback_data=f"setclass|{t['id']}")])
        kb_rows.append([InlineKeyboardButton(text=f"✏️ Предмет {t['id']}", callback_data=f"setsubj|{t['id']}")])

    kb_rows.append([InlineKeyboardButton(text="🔙 Назад", callback_data="admin_back")])

    await message.answer("\n".join(lines), reply_markup=InlineKeyboardMarkup(inline_keyboard=kb_rows))


@dp.callback_query(F.data == "admin_broadcast")
async def admin_broadcast(cb: types.CallbackQuery, state: FSMContext):
    if not is_admin(cb.from_user.id):
        await safe_answer(cb, "⛔ Нет доступа", show_alert=True)
        return
    await state.set_state(AdminStates.waiting_for_broadcast)
    await safe_answer(cb)
    await cb.message.answer("Введите сообщение для рассылки ВСЕМ подтвержденным учителям:")


@dp.message(AdminStates.waiting_for_broadcast)
async def st_wait_broadcast(message: types.Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return

    text = (message.text or "").strip()
    if not text:
        await message.answer("Введите непустое сообщение.")
        return

    await state.clear()

    teachers = get_approved_teachers()
    sent = 0
    failed = 0

    for t in teachers:
        try:
            await bot.send_message(int(t["telegram_id"]), f"📢 Сообщение администрации:\n\n{text}")
            sent += 1
            await asyncio.sleep(0.05)
        except Exception as e:
            failed += 1
            logger.error(f"broadcast error to {t.get('telegram_id')}: {e}")

    await message.answer(f"✅ Рассылка завершена. Отправлено: {sent}, ошибок: {failed}. /admin")


@dp.callback_query(F.data.startswith("appr|"))
async def cb_approve(cb: types.CallbackQuery):
    if not is_admin(cb.from_user.id):
        await safe_answer(cb, "⛔ Нет доступа", show_alert=True)
        return
    teacher_id = safe_int(cb.data.split("|")[1])
    approve_teacher_db(teacher_id)
    await safe_answer(cb, "✅ Подтверждено")


@dp.callback_query(F.data.startswith("decl|"))
async def cb_decline(cb: types.CallbackQuery):
    if not is_admin(cb.from_user.id):
        await safe_answer(cb, "⛔ Нет доступа", show_alert=True)
        return
    teacher_id = safe_int(cb.data.split("|")[1])
    decline_teacher_db(teacher_id)
    await safe_answer(cb, "🗑 Удалено")


@dp.callback_query(F.data.startswith("setclass|"))
async def cb_setclass(cb: types.CallbackQuery, state: FSMContext):
    if not is_admin(cb.from_user.id):
        await safe_answer(cb, "⛔ Нет доступа", show_alert=True)
        return
    teacher_id = safe_int(cb.data.split("|")[1])
    await state.update_data(edit_teacher_id=teacher_id, edit_field="class_name")
    await state.set_state(AdminStates.waiting_for_class)
    await safe_answer(cb)
    await cb.message.answer(f"Введите класс для учителя id <code>{teacher_id}</code> (например 7А):")


@dp.callback_query(F.data.startswith("setsubj|"))
async def cb_setsubj(cb: types.CallbackQuery, state: FSMContext):
    if not is_admin(cb.from_user.id):
        await safe_answer(cb, "⛔ Нет доступа", show_alert=True)
        return
    teacher_id = safe_int(cb.data.split("|")[1])
    await state.update_data(edit_teacher_id=teacher_id, edit_field="subject")
    await state.set_state(AdminStates.waiting_for_subject)
    await safe_answer(cb)
    await cb.message.answer(f"Введите предмет для учителя id <code>{teacher_id}</code> (например Информатика):")


@dp.message(AdminStates.waiting_for_class)
async def st_wait_class(message: types.Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    data = await state.get_data()
    teacher_id = safe_int(data.get("edit_teacher_id"))
    val = (message.text or "").strip()
    if not val:
        await message.answer("Введите значение класса.")
        return
    update_teacher_field(teacher_id, "class_name", val)
    await state.clear()
    await message.answer("✅ Класс обновлен. /admin")


@dp.message(AdminStates.waiting_for_subject)
async def st_wait_subject(message: types.Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    data = await state.get_data()
    teacher_id = safe_int(data.get("edit_teacher_id"))
    val = (message.text or "").strip()
    if not val:
        await message.answer("Введите значение предмета.")
        return
    update_teacher_field(teacher_id, "subject", val)
    await state.clear()
    await message.answer("✅ Предмет обновлен. /admin")


# =========================
# CALLBACKS (report) + анти-лаг
# =========================
@dp.callback_query(F.data == "noop")
async def cb_noop(cb: types.CallbackQuery):
    await safe_answer(cb)


@dp.callback_query(F.data == "manual_report_start")
async def cb_manual_report_start(cb: types.CallbackQuery):
    lock = USER_LOCKS[cb.from_user.id]
    if lock.locked():
        await safe_answer(cb, "⏳ Подождите…")
        return
    async with lock:
        await safe_answer(cb)
        await show_report(cb.from_user.id, cb.from_user.id)


@dp.callback_query(F.data.startswith(("edit|", "set|", "delta|", "back|", "refresh|", "submit|")))
async def cb_report_router(cb: types.CallbackQuery):
    key = cb_seen_key(cb)
    if key in SEEN_CALLBACKS:
        await safe_answer(cb)
        return
    SEEN_CALLBACKS.add(key)
    if len(SEEN_CALLBACKS) > 50000:
        SEEN_CALLBACKS.clear()

    lock = USER_LOCKS[cb.from_user.id]
    if lock.locked():
        await safe_answer(cb, "⏳ Подождите…")
        return

    async with lock:
        await safe_answer(cb)

        t = get_teacher_by_telegram(cb.from_user.id)
        if not t or safe_int(t.get("is_approved")) != 1:
            await safe_answer(cb, "⛔ Нет доступа", show_alert=True)
            return

        parts = cb.data.split("|")
        action = parts[0]

        if action == "edit":
            field = parts[1]
            report_id = safe_int(parts[2])

            conn = get_connection()
            cur = conn.cursor()
            cur.execute("SELECT * FROM daily_reports WHERE id = ?", (report_id,))
            rep_row = cur.fetchone()
            conn.close()

            if not rep_row:
                await safe_answer(cb, "❌ Отчет не найден", show_alert=True)
                return

            rep = dict(rep_row)
            if safe_int(rep.get("is_submitted")) == 1:
                await safe_answer(cb, "✅ Уже отправлено", show_alert=True)
                return

            current = safe_int(rep.get(field))
            await safe_edit_text(
                cb.message,
                f"Выберите число.\nПоле: <b>{field}</b>\nТекущее: <b>{current}</b>",
                number_kb(field, current, report_id),
            )
            return

        if action == "set":
            field = parts[1]
            value = safe_int(parts[2])
            report_id = safe_int(parts[3])

            conn = get_connection()
            cur = conn.cursor()
            cur.execute("SELECT * FROM daily_reports WHERE id = ?", (report_id,))
            rep_row = cur.fetchone()
            conn.close()

            if not rep_row:
                await safe_answer(cb, "❌ Отчет не найден", show_alert=True)
                return

            rep = dict(rep_row)
            if safe_int(rep.get("is_submitted")) == 1:
                await safe_answer(cb, "✅ Уже отправлено", show_alert=True)
                return

            set_report_value(report_id, field, value)

            rep2 = get_connection()
            rep2.close()

            rep_new = None
            conn = get_connection()
            cur = conn.cursor()
            cur.execute("SELECT * FROM daily_reports WHERE id = ?", (report_id,))
            rr = cur.fetchone()
            conn.close()
            if rr:
                rep_new = dict(rr)

            cur_val = safe_int((rep_new or rep).get(field))
            await safe_edit_text(
                cb.message,
                f"Выберите число.\nПоле: <b>{field}</b>\nТекущее: <b>{cur_val}</b>",
                number_kb(field, cur_val, report_id),
            )
            return

        if action == "delta":
            field = parts[1]
            delta = safe_int(parts[2])
            report_id = safe_int(parts[3])

            conn = get_connection()
            cur = conn.cursor()
            cur.execute("SELECT * FROM daily_reports WHERE id = ?", (report_id,))
            rep_row = cur.fetchone()
            conn.close()

            if not rep_row:
                await safe_answer(cb, "❌ Отчет не найден", show_alert=True)
                return

            rep = dict(rep_row)
            if safe_int(rep.get("is_submitted")) == 1:
                await safe_answer(cb, "✅ Уже отправлено", show_alert=True)
                return

            new_val = safe_int(rep.get(field)) + delta
            set_report_value(report_id, field, new_val)

            rep_new = get_report_by_teacher_id(int(rep["teacher_id"]), rep["report_date"])  # type: ignore
            # Если не нашли — просто перечитаем по id
            if rep_new is None:
                conn = get_connection()
                cur = conn.cursor()
                cur.execute("SELECT * FROM daily_reports WHERE id = ?", (report_id,))
                rr = cur.fetchone()
                conn.close()
                rep_new = dict(rr) if rr else rep

            cur_val = safe_int(rep_new.get(field))
            await safe_edit_text(
                cb.message,
                f"Выберите число.\nПоле: <b>{field}</b>\nТекущее: <b>{cur_val}</b>",
                number_kb(field, cur_val, report_id),
            )
            return

        if action == "back":
            report_id = safe_int(parts[1])
            conn = get_connection()
            cur = conn.cursor()
            cur.execute("SELECT * FROM daily_reports WHERE id = ?", (report_id,))
            rep_row = cur.fetchone()
            conn.close()
            if not rep_row:
                await safe_answer(cb, "❌ Отчет не найден", show_alert=True)
                return
            rep = dict(rep_row)
            await safe_edit_text(cb.message, report_text(rep), report_kb(rep))
            return

        if action == "refresh":
            report_id = safe_int(parts[1])
            conn = get_connection()
            cur = conn.cursor()
            cur.execute("SELECT * FROM daily_reports WHERE id = ?", (report_id,))
            rep_row = cur.fetchone()
            conn.close()
            if not rep_row:
                await safe_answer(cb, "❌ Отчет не найден", show_alert=True)
                return
            rep = dict(rep_row)
            await safe_edit_text(cb.message, report_text(rep), report_kb(rep))
            return

        if action == "submit":
            report_id = safe_int(parts[1])
            conn = get_connection()
            cur = conn.cursor()
            cur.execute("SELECT * FROM daily_reports WHERE id = ?", (report_id,))
            rep_row = cur.fetchone()
            conn.close()
            if not rep_row:
                await safe_answer(cb, "❌ Отчет не найден", show_alert=True)
                return

            rep = dict(rep_row)
            if safe_int(rep.get("is_submitted")) == 1:
                await safe_answer(cb, "✅ Уже отправлено", show_alert=True)
                return

            ok = submit_report(report_id)
            if not ok:
                await safe_answer(cb, "❌ Не удалось отправить", show_alert=True)
                return

            await safe_edit_text(cb.message, "✅ Отчет успешно отправлен!")
            return


# =========================
# REGISTRATION (секретный код)
# =========================
@dp.message()
async def catch_secret_code(message: types.Message):
    status = get_teacher_status(message.from_user.id)
    if status != "guest":
        return

    txt = (message.text or "").strip()
    if txt != SECRET_CODE:
        return

    full_name = (message.from_user.full_name or "").strip() or "Учитель"
    t = create_teacher_if_missing(message.from_user.id, full_name)
    await message.answer("✅ Заявка создана. Ожидайте подтверждения администратора.")

    for admin_id in ADMIN_IDS:
        try:
            await bot.send_message(
                admin_id,
                "🆕 Новая заявка учителя:\n"
                f"ФИО: <b>{t.get('full_name')}</b>\n"
                f"telegram_id: <code>{t.get('telegram_id')}</code>\n"
                f"teacher_id: <code>{t.get('id')}</code>\n\n"
                "Зайдите в /admin → 👥 Ожидают и подтвердите.",
            )
        except Exception as e:
            logger.error(f"notify admin new teacher error: {e}")


# =========================
# MAIN
# =========================
async def main():
    init_db()
    logger.info("==================================================")
    logger.info("🤖 Бот учета посещаемости запускается...")
    logger.info(f"👑 Администраторы: {ADMIN_IDS}")
    logger.info("🔑 SECRET_CODE: (скрыт)")
    logger.info(f"📨 REPORT_CHAT_ID: {REPORT_CHAT_ID}")
    logger.info("==================================================")

    asyncio.create_task(schedule_loop())
    await dp.start_polling(bot)


if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        pass
