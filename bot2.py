import asyncio
import logging
import re
import sqlite3
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime, date
from io import BytesIO
from typing import Optional, Dict, Any, List

from aiogram import Bot, Dispatcher, types, F
from aiogram.client.default import DefaultBotProperties
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton, BufferedInputFile

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# =========================
# SETTINGS
# =========================
BOT_TOKEN = "8524573448:AAFHW_KFr8_M2iJ5tp5rZRSPkj_GVjM-V34"
DATABASE_FILE = "attendance.db"

# Замените на свой ID
ADMIN_IDS: List[int] = [7233585816]
SECRET_CODE = "учитель2026"

MAX_STUDENTS = 32

# -100... id приватной группы (бот должен быть в группе и иметь право писать)
REPORT_CHAT_ID: Optional[int] = -1003756818645

LOG_LEVEL = "INFO"

# Значения по умолчанию (если в БД нет настроек)
DEFAULT_TOTAL_SHIFT_1 = 1067
DEFAULT_TOTAL_SHIFT_2 = 256


# =========================
# LOGGING
# =========================
logging.basicConfig(
    level=getattr(logging, LOG_LEVEL, logging.INFO),
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    handlers=[logging.FileHandler("bot.log", encoding="utf-8"), logging.StreamHandler()],
)
logger = logging.getLogger("attendance_bot")


# =========================
# BOT / DISPATCHER
# =========================
bot = Bot(token=BOT_TOKEN, default=DefaultBotProperties(parse_mode="HTML"))
dp = Dispatcher(storage=MemoryStorage())

# анти-лаг: один юзер = одна операция callback за раз
USER_LOCKS: Dict[int, asyncio.Lock] = defaultdict(asyncio.Lock)
SEEN_CALLBACKS: set[str] = set()


# =========================
# FSM
# =========================
class AdminStates(StatesGroup):
    waiting_for_subject = State()
    waiting_for_class = State()
    waiting_for_broadcast = State()
    waiting_for_search = State()
    waiting_for_setting_val = State() # Для ввода числа учеников


# =========================
# DB
# =========================
def get_connection() -> sqlite3.Connection:
    conn = sqlite3.connect(DATABASE_FILE, timeout=30)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    conn.execute("PRAGMA journal_mode = WAL")
    return conn

def ensure_column(conn: sqlite3.Connection, table: str, column: str, col_def: str):
    cur = conn.cursor()
    cur.execute(f"PRAGMA table_info({table})")
    cols = {row[1] for row in cur.fetchall()}
    if column not in cols:
        cur.execute(f"ALTER TABLE {table} ADD COLUMN {column} {col_def}")
        conn.commit()

def init_db():
    conn = get_connection()
    cur = conn.cursor()

    # Таблица учителей
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS teachers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            telegram_id INTEGER UNIQUE NOT NULL,
            username TEXT,
            full_name TEXT NOT NULL,
            subject TEXT,
            class_name TEXT,
            is_approved INTEGER DEFAULT 0,
            registered_at TEXT DEFAULT (datetime('now'))
        )
        """
    )
    # Таблица отчетов
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS daily_reports (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            teacher_id INTEGER NOT NULL,
            report_date TEXT NOT NULL,

            sick INTEGER DEFAULT 0,
            sanatorium INTEGER DEFAULT 0,
            parent_statement INTEGER DEFAULT 0,
            competition INTEGER DEFAULT 0,
            absent_without_reason INTEGER DEFAULT 0,

            is_submitted INTEGER DEFAULT 0,
            submitted_at TEXT,

            FOREIGN KEY (teacher_id) REFERENCES teachers (id) ON DELETE CASCADE,
            UNIQUE(teacher_id, report_date)
        )
        """
    )
    # Таблица настроек (ключ-значение)
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value INTEGER
        )
        """
    )
    # Инициализация дефолтных настроек, если их нет
    cur.execute("INSERT OR IGNORE INTO settings (key, value) VALUES (?, ?)", ("total_1", DEFAULT_TOTAL_SHIFT_1))
    cur.execute("INSERT OR IGNORE INTO settings (key, value) VALUES (?, ?)", ("total_2", DEFAULT_TOTAL_SHIFT_2))

    # Обновление структуры (если старая база)
    ensure_column(conn, "teachers", "subject", "TEXT")
    ensure_column(conn, "teachers", "class_name", "TEXT")
    ensure_column(conn, "teachers", "username", "TEXT")

    conn.commit()
    conn.close()
    logger.info("✅ База данных проверена/инициализирована")


# =========================
# DB HELPERS
# =========================
def update_username_if_changed(telegram_id: int, new_username: Optional[str]):
    """Тихо обновляет username, если он изменился или отсутствует"""
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT username FROM teachers WHERE telegram_id = ?", (telegram_id,))
    row = cur.fetchone()
    
    if row:
        current_db_username = row["username"]
        # Если отличается от реального — обновляем
        if current_db_username != new_username:
            cur.execute("UPDATE teachers SET username = ? WHERE telegram_id = ?", (new_username, telegram_id))
            conn.commit()
    conn.close()

def get_setting(key: str, default: int) -> int:
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT value FROM settings WHERE key = ?", (key,))
    row = cur.fetchone()
    conn.close()
    return int(row["value"]) if row else default

def set_setting(key: str, value: int):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("INSERT INTO settings(key, value) VALUES(?, ?) ON CONFLICT(key) DO UPDATE SET value=excluded.value", (key, value))
    conn.commit()
    conn.close()

def get_total_students_for_shift(second_shift: bool) -> int:
    key = "total_2" if second_shift else "total_1"
    default = DEFAULT_TOTAL_SHIFT_2 if second_shift else DEFAULT_TOTAL_SHIFT_1
    return get_setting(key, default)

def get_teacher_by_telegram(telegram_id: int) -> Optional[Dict[str, Any]]:
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT * FROM teachers WHERE telegram_id = ?", (telegram_id,))
    row = cur.fetchone()
    conn.close()
    return dict(row) if row else None

def get_teacher_by_id(t_id: int) -> Optional[Dict[str, Any]]:
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT * FROM teachers WHERE id = ?", (t_id,))
    row = cur.fetchone()
    conn.close()
    return dict(row) if row else None

def create_teacher_if_missing(telegram_id: int, full_name: str, username: Optional[str]) -> Dict[str, Any]:
    t = get_teacher_by_telegram(telegram_id)
    if t:
        return t
    conn = get_connection()
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO teachers(telegram_id, full_name, username, is_approved) VALUES (?, ?, ?, 0)",
        (telegram_id, full_name, username),
    )
    conn.commit()
    conn.close()
    return get_teacher_by_telegram(telegram_id) or {}

def approve_teacher_db(teacher_id: int):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("UPDATE teachers SET is_approved = 1 WHERE id = ?", (teacher_id,))
    conn.commit()
    conn.close()

def decline_teacher_db(teacher_id: int):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM teachers WHERE id = ?", (teacher_id,))
    conn.commit()
    conn.close()

def update_teacher_field(teacher_id: int, field: str, value: str):
    if field not in {"subject", "class_name", "full_name", "username"}:
        return
    conn = get_connection()
    cur = conn.cursor()
    cur.execute(f"UPDATE teachers SET {field} = ? WHERE id = ?", (value, teacher_id))
    conn.commit()
    conn.close()

def get_teachers_by_status(approved: bool) -> List[Dict[str, Any]]:
    conn = get_connection()
    cur = conn.cursor()
    is_appr = 1 if approved else 0
    cur.execute(
        "SELECT * FROM teachers WHERE is_approved = ? ORDER BY class_name, full_name",
        (is_appr,)
    )
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return rows

def search_teachers(q: str) -> List[Dict[str, Any]]:
    q = (q or "").strip()
    if not q:
        return []
    like = f"%{q.lower()}%"
    conn = get_connection()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT * FROM teachers
        WHERE lower(full_name) LIKE ?
           OR lower(class_name) LIKE ?
           OR lower(username) LIKE ?
           OR CAST(telegram_id AS TEXT) LIKE ?
        ORDER BY is_approved DESC, class_name
        """,
        (like, like, like, f"%{q}%"),
    )
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return rows

def get_report_by_teacher_id(teacher_id: int, report_date_iso: str) -> Optional[Dict[str, Any]]:
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT * FROM daily_reports WHERE teacher_id = ? AND report_date = ?", (teacher_id, report_date_iso))
    row = cur.fetchone()
    conn.close()
    return dict(row) if row else None

def create_or_get_report(teacher_id: int, report_date_iso: str) -> Dict[str, Any]:
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT * FROM daily_reports WHERE teacher_id = ? AND report_date = ?", (teacher_id, report_date_iso))
    row = cur.fetchone()
    if row:
        conn.close()
        return dict(row)
    cur.execute("INSERT INTO daily_reports(teacher_id, report_date) VALUES(?, ?)", (teacher_id, report_date_iso))
    conn.commit()
    cur.execute("SELECT * FROM daily_reports WHERE teacher_id = ? AND report_date = ?", (teacher_id, report_date_iso))
    row2 = cur.fetchone()
    conn.close()
    return dict(row2) if row2 else {}

def set_report_value(report_id: int, field: str, value: int) -> bool:
    if field not in {"sick", "sanatorium", "parent_statement", "competition", "absent_without_reason"}:
        return False
    value = max(0, min(int(MAX_STUDENTS), safe_int(value)))
    conn = get_connection()
    cur = conn.cursor()
    cur.execute(f"UPDATE daily_reports SET {field} = ? WHERE id = ?", (value, report_id))
    conn.commit()
    ok = cur.rowcount > 0
    conn.close()
    return ok

def submit_report(report_id: int) -> bool:
    conn = get_connection()
    cur = conn.cursor()
    cur.execute(
        "UPDATE daily_reports SET is_submitted = 1, submitted_at = datetime('now') WHERE id = ? AND is_submitted = 0",
        (report_id,),
    )
    conn.commit()
    ok = cur.rowcount > 0
    conn.close()
    return ok

def create_today_reports_for_all_approved():
    d = today_iso()
    for t in get_teachers_by_status(approved=True):
        create_or_get_report(int(t["id"]), d)


# =========================
# GENERAL HELPERS
# =========================
def today_iso() -> str:
    return date.today().isoformat()

def safe_int(x: Any, default: int = 0) -> int:
    try:
        return int(x)
    except Exception:
        return default

def is_admin(user_id: int) -> bool:
    return user_id in ADMIN_IDS

def grade_from_class_name(class_name: Optional[str]) -> Optional[int]:
    if not class_name: return None
    m = re.match(r"^\s*(\d+)", str(class_name))
    return int(m.group(1)) if m else None

def is_second_shift_class(class_name: Optional[str]) -> bool:
    # 2 смена = 3* и 6*
    g = grade_from_class_name(class_name)
    return g in {3, 6}

def get_shift_teachers(second_shift: bool) -> List[Dict[str, Any]]:
    rows = get_teachers_by_status(approved=True)
    return [t for t in rows if is_second_shift_class(t.get("class_name")) == second_shift]

async def safe_answer(cb: types.CallbackQuery, text: Optional[str] = None, show_alert: bool = False):
    try:
        await cb.answer(text=text, show_alert=show_alert)
    except Exception:
        pass

async def safe_edit_text(msg: types.Message, text: str, reply_markup: Optional[InlineKeyboardMarkup] = None):
    try:
        await msg.edit_text(text, reply_markup=reply_markup)
    except Exception:
        try:
            await msg.answer(text, reply_markup=reply_markup)
        except Exception:
            pass

def format_teacher_label(t: Dict[str, Any]) -> str:
    cls = t.get("class_name") or "?"
    name = t.get("full_name")
    return f"{cls} — {name}"


# =========================
# KEYBOARDS
# =========================
def admin_main_kb() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(text="👥 Ожидают", callback_data="admin_pending"),
                InlineKeyboardButton(text="✅ Подтвержденные", callback_data="admin_approved"),
            ],
            [
                InlineKeyboardButton(text="🔍 Поиск", callback_data="admin_search"),
                InlineKeyboardButton(text="⚙️ Настройки", callback_data="admin_settings"),
            ],
            [
                InlineKeyboardButton(text="📢 Рассылка", callback_data="admin_broadcast"),
                InlineKeyboardButton(text="🔄 Обновить", callback_data="admin_refresh"),
            ],
        ]
    )

def settings_kb() -> InlineKeyboardMarkup:
    t1 = get_setting("total_1", DEFAULT_TOTAL_SHIFT_1)
    t2 = get_setting("total_2", DEFAULT_TOTAL_SHIFT_2)
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text=f"1 смена (сейчас: {t1})", callback_data="set_total|1")],
            [InlineKeyboardButton(text=f"2 смена (сейчас: {t2})", callback_data="set_total|2")],
            [InlineKeyboardButton(text="🔙 Назад", callback_data="admin_back")],
        ]
    )

def back_kb(cb: str = "admin_back") -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text="🔙 Назад", callback_data=cb)]])

def teacher_manage_kb(teacher_id: int) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(text="✏️ Класс", callback_data=f"setclass|{teacher_id}"),
                InlineKeyboardButton(text="✏️ Предмет", callback_data=f"setsubj|{teacher_id}"),
            ],
            [InlineKeyboardButton(text="❌ Удалить учителя", callback_data=f"decl|{teacher_id}")],
            [InlineKeyboardButton(text="🔙 К списку", callback_data="admin_approved")],
        ]
    )

def teacher_pending_kb(teacher_id: int) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(text="✅ Принять", callback_data=f"appr|{teacher_id}"),
                InlineKeyboardButton(text="❌ Отклонить", callback_data=f"decl|{teacher_id}"),
            ],
            [
                InlineKeyboardButton(text="✏️ Класс", callback_data=f"setclass|{teacher_id}"),
                InlineKeyboardButton(text="✏️ Предмет", callback_data=f"setsubj|{teacher_id}"),
            ],
            [InlineKeyboardButton(text="🔙 К списку", callback_data="admin_pending")],
        ]
    )

def report_kb(report: Dict[str, Any]) -> InlineKeyboardMarkup:
    rid = int(report["id"])
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="🤒 Болеют", callback_data=f"edit|sick|{rid}")],
            [InlineKeyboardButton(text="🏥 На оздоровлении", callback_data=f"edit|sanatorium|{rid}")],
            [InlineKeyboardButton(text="📄 По заявлению", callback_data=f"edit|parent_statement|{rid}")],
            [InlineKeyboardButton(text="🏆 На соревнованиях", callback_data=f"edit|competition|{rid}")],
            [InlineKeyboardButton(text="❌ Без причины", callback_data=f"edit|absent_without_reason|{rid}")],
            [
                InlineKeyboardButton(text="📤 Отправить отчет", callback_data=f"submit|{rid}"),
                InlineKeyboardButton(text="🔄 Обновить", callback_data=f"refresh|{rid}"),
            ],
        ]
    )

def number_kb(field: str, current: int, report_id: int) -> InlineKeyboardMarkup:
    max_n = MAX_STUDENTS
    numbers = list(range(0, max_n + 1))
    rows = []
    row = []
    for n in numbers:
        txt = f"✅ {n}" if n == current else str(n)
        row.append(InlineKeyboardButton(text=txt, callback_data=f"set|{field}|{n}|{report_id}"))
        if len(row) == 6:
            rows.append(row)
            row = []
    if row: rows.append(row)
    rows.append([
        InlineKeyboardButton(text=f"Сейчас: {current}", callback_data="noop"),
        InlineKeyboardButton(text="+1", callback_data=f"delta|{field}|1|{report_id}"),
        InlineKeyboardButton(text="+5", callback_data=f"delta|{field}|5|{report_id}"),
    ])
    rows.append([
        InlineKeyboardButton(text="Сброс 0", callback_data=f"set|{field}|0|{report_id}"),
        InlineKeyboardButton(text="🔙 Назад", callback_data=f"back|{report_id}"),
    ])
    return InlineKeyboardMarkup(inline_keyboard=rows)


# =========================
# EXCEL GENERATION
# =========================
def build_shift_excel(report_date_iso: str, second_shift: bool, teachers: List[Dict[str, Any]]) -> bytes:
    title = "2 смена (3* и 6*)" if second_shift else "1 смена (кроме 3* и 6*)"
    
    # Получаем актуальное количество учеников из настроек
    total_students_count = get_total_students_for_shift(second_shift)

    wb = Workbook()
    ws = wb.active
    ws.title = "Отчеты"

    # Стили
    header_fill = PatternFill("solid", fgColor="E7EEF7")
    total_fill = PatternFill("solid", fgColor="FFF2CC")
    percent_fill = PatternFill("solid", fgColor="E2F0D9")
    present_fill = PatternFill("solid", fgColor="C6E0B4") # Зеленый для присутствующих

    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="A0A0A0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Заголовок
    ws["A1"] = f"Отчет за {report_date_iso} — {title}"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:G1")
    ws.row_dimensions[1].height = 24

    headers = [
        "Класс", "Статус", "Болеют", "Санаторий",
        "По заявлению", "На соревнованиях", "Без причины"
    ]
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=2, column=col)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[2].height = 34

    # Переменные для сумм
    total_sick = total_san = total_parent = total_comp = total_no = 0
    submitted_cnt = 0
    pending_cnt = 0

    # Заполнение данными
    row_idx = 3
    for t in teachers:
        rep = get_report_by_teacher_id(int(t["id"]), report_date_iso)
        cls = t.get("class_name") or "—"
        
        if rep and safe_int(rep.get("is_submitted")) == 1:
            status = "Сдан"
            submitted_cnt += 1
            sick = safe_int(rep.get("sick"))
            san = safe_int(rep.get("sanatorium"))
            par = safe_int(rep.get("parent_statement"))
            comp = safe_int(rep.get("competition"))
            no = safe_int(rep.get("absent_without_reason"))
            
            total_sick += sick
            total_san += san
            total_parent += par
            total_comp += comp
            total_no += no
        else:
            status = "Не сдан"
            pending_cnt += 1
            sick = san = par = comp = no = 0

        ws.append([cls, status, sick, san, par, comp, no])
        
        for c in range(1, 8):
            cell = ws.cell(row=row_idx, column=c)
            cell.border = border
            cell.alignment = left if c == 1 else center
        row_idx += 1

    # ИТОГИ (Отсутствующие)
    ws.append(["ИТОГО (Отсутствует)", f"Сдано: {submitted_cnt}", total_sick, total_san, total_parent, total_comp, total_no])
    total_row = row_idx
    
    for c in range(1, 8):
        cell = ws.cell(row=total_row, column=c)
        cell.font = bold
        cell.fill = total_fill
        cell.border = border
        cell.alignment = center if c >= 2 else left

    # ПРОЦЕНТЫ (Отсутствующие)
    denom = total_students_count if total_students_count > 0 else 1
    ws.append([
        "ПРОЦЕНТ (Отсутствует)", f"от {total_students_count}",
        total_sick / denom, total_san / denom, total_parent / denom,
        total_comp / denom, total_no / denom
    ])
    percent_row = total_row + 1
    
    for c in range(1, 8):
        cell = ws.cell(row=percent_row, column=c)
        cell.font = bold
        cell.fill = percent_fill
        cell.border = border
        cell.alignment = center if c >= 2 else left
    for c in range(3, 8):
        ws.cell(row=percent_row, column=c).number_format = "0.00%"

    # === НОВАЯ СЕКЦИЯ: ПРИСУТСТВУЮЩИЕ ===
    total_absent_sum = total_sick + total_san + total_parent + total_comp + total_no
    total_present = max(0, total_students_count - total_absent_sum)
    present_pct = total_present / denom

    # Строка количества присутствующих
    ws.append(["ПРИСУТСТВУЕТ (Чел.)", "", total_present, "", "", "", ""])
    present_row_num = percent_row + 1
    
    # Объединяем ячейки для красивого вида
    ws.merge_cells(start_row=present_row_num, start_column=3, end_row=present_row_num, end_column=7)
    
    # Стилизация строки присутствующих
    ws.cell(row=present_row_num, column=1).font = bold
    ws.cell(row=present_row_num, column=1).fill = present_fill
    ws.cell(row=present_row_num, column=1).border = border
    
    val_cell = ws.cell(row=present_row_num, column=3)
    val_cell.font = bold
    val_cell.fill = present_fill
    val_cell.alignment = center
    val_cell.border = border

    # Строка процента присутствующих
    ws.append(["ПРИСУТСТВУЕТ (%)", "", present_pct, "", "", "", ""])
    present_pct_row = present_row_num + 1
    
    ws.merge_cells(start_row=present_pct_row, start_column=3, end_row=present_pct_row, end_column=7)
    
    ws.cell(row=present_pct_row, column=1).font = bold
    ws.cell(row=present_pct_row, column=1).fill = present_fill
    ws.cell(row=present_pct_row, column=1).border = border
    
    pct_cell = ws.cell(row=present_pct_row, column=3)
    pct_cell.font = bold
    pct_cell.fill = present_fill
    pct_cell.alignment = center
    pct_cell.border = border
    pct_cell.number_format = "0.00%"

    # Ширина колонок
    col_widths = [10, 22, 12, 12, 22, 18, 24]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A3"
    
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()

async def send_shift_excel_to_chat(report_date_iso: str, second_shift: bool, teachers: List[Dict[str, Any]]):
    file_bytes = build_shift_excel(report_date_iso, second_shift, teachers)
    shift_label = "2смена" if second_shift else "1смена"
    filename = f"Отчеты_{report_date_iso}_{shift_label}.xlsx"
    doc = BufferedInputFile(file_bytes, filename=filename)

    if REPORT_CHAT_ID:
        try:
            await bot.send_document(REPORT_CHAT_ID, doc, caption=f"📎 Итог {report_date_iso} ({shift_label})")
        except Exception as e:
            logger.error(f"send_shift_excel error: {e}")
        return

    for admin_id in ADMIN_IDS:
        try:
            await bot.send_document(admin_id, doc, caption=f"📎 Итог {report_date_iso} ({shift_label})")
        except:
            pass


# =========================
# SCHEDULE LOGIC
# =========================
@dataclass(frozen=True)
class ScheduleEvent:
    hhmm: str
    second_shift: bool
    label: str
    is_summary: bool

SCHEDULE: List[ScheduleEvent] = [
    ScheduleEvent("07:50", False, "Напоминание 1/4", False),
    ScheduleEvent("08:10", False, "Напоминание 2/4", False),
    ScheduleEvent("08:20", False, "Напоминание 3/4", False),
    ScheduleEvent("08:30", False, "Напоминание 4/4", False),
    ScheduleEvent("08:40", False, "Итог", True),

    ScheduleEvent("12:50", True, "Напоминание 1/4", False),
    ScheduleEvent("13:10", True, "Напоминание 2/4", False),
    ScheduleEvent("13:20", True, "Напоминание 3/4", False),
    ScheduleEvent("13:30", True, "Напоминание 4/4", False),
    ScheduleEvent("13:40", True, "Итог", True),
]

_EXECUTED: set[str] = set()

async def notify_teacher_fill(telegram_id: int, report_date_iso: str, label: str) -> bool:
    try:
        await bot.send_message(
            telegram_id,
            f"🔔 <b>{label}</b>\nПожалуйста, заполните отчет за <b>{report_date_iso}</b>.",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text="📊 Заполнить отчет", callback_data="manual_report_start")]])
        )
        return True
    except Exception:
        return False

async def schedule_loop():
    while True:
        try:
            now = datetime.now()
            d_iso = now.date().isoformat()
            hhmm = now.strftime("%H:%M")

            for ev in SCHEDULE:
                if ev.hhmm != hhmm:
                    continue

                key = f"{d_iso}|{hhmm}|{ev.second_shift}"
                if key in _EXECUTED:
                    continue
                _EXECUTED.add(key)

                create_today_reports_for_all_approved()
                teachers = get_shift_teachers(ev.second_shift)

                if ev.is_summary:
                    await send_shift_excel_to_chat(d_iso, ev.second_shift, teachers)
                else:
                    sent_count = 0
                    for t in teachers:
                        rep = get_report_by_teacher_id(int(t["id"]), d_iso)
                        # Если уже сдал - не беспокоим
                        if rep and safe_int(rep.get("is_submitted")) == 1:
                            continue
                        
                        success = await notify_teacher_fill(int(t["telegram_id"]), d_iso, ev.label)
                        if success:
                            sent_count += 1
                        await asyncio.sleep(0.05)
                    
                    # Уведомление администратору о рассылке
                    for admin_id in ADMIN_IDS:
                        try:
                            await bot.send_message(
                                admin_id, 
                                f"📢 Авто-рассылка '{ev.label}' завершена.\n"
                                f"Получателей (кто еще не сдал): <b>{sent_count}</b>"
                            )
                        except: pass

            if len(_EXECUTED) > 5000:
                _EXECUTED.clear()

            await asyncio.sleep(2)
        except Exception as e:
            logger.error(f"schedule_loop error: {e}")
            await asyncio.sleep(5)


# =========================
# ADMIN UI & HANDLERS
# =========================
@dp.message(Command("admin"))
async def cmd_admin(message: types.Message, state: FSMContext):
    if not is_admin(message.from_user.id): return
    await state.clear()
    await admin_render_main(message)

async def admin_render_main(msg: types.Message):
    create_today_reports_for_all_approved()
    today = today_iso()
    
    teachers_appr = get_teachers_by_status(True)
    teachers_wait = get_teachers_by_status(False)
    
    conn = get_connection()
    cur = conn.cursor()
    cur.execute(
        "SELECT COUNT(*) as tot, SUM(CASE WHEN is_submitted=1 THEN 1 ELSE 0 END) as sub FROM daily_reports WHERE report_date=?",
        (today,)
    )
    st = cur.fetchone()
    conn.close()
    
    sub = st['sub'] or 0
    tot = st['tot'] or 0

    text = (
        "👑 <b>Админ-панель</b>\n\n"
        f"📅 <b>{today}</b>\n"
        f"✅ Учителей: <b>{len(teachers_appr)}</b>\n"
        f"⏳ Заявок: <b>{len(teachers_wait)}</b>\n"
        f"📊 Сдано отчетов: <b>{sub} / {tot}</b>"
    )
    try:
        await msg.edit_text(text, reply_markup=admin_main_kb())
    except:
        await msg.answer(text, reply_markup=admin_main_kb())

@dp.callback_query(F.data == "admin_back")
async def cb_admin_back(cb: types.CallbackQuery, state: FSMContext):
    if not is_admin(cb.from_user.id): return
    await state.clear()
    await admin_render_main(cb.message)
    await safe_answer(cb)

@dp.callback_query(F.data == "admin_refresh")
async def cb_admin_refresh(cb: types.CallbackQuery):
    if not is_admin(cb.from_user.id): return
    await admin_render_main(cb.message)
    await safe_answer(cb, "Обновлено")

# --- Настройки (Кол-во учеников) ---
@dp.callback_query(F.data == "admin_settings")
async def cb_admin_settings(cb: types.CallbackQuery):
    if not is_admin(cb.from_user.id): return
    await cb.message.edit_text("⚙️ <b>Настройки количества учащихся</b>\nВыберите смену для изменения:", reply_markup=settings_kb())

@dp.callback_query(F.data.startswith("set_total|"))
async def cb_set_total(cb: types.CallbackQuery, state: FSMContext):
    shift_num = cb.data.split("|")[1]
    await state.update_data(shift_num=shift_num)
    await state.set_state(AdminStates.waiting_for_setting_val)
    await cb.message.answer(f"Введите новое общее количество учащихся для <b>{shift_num} смены</b>:")
    await safe_answer(cb)

@dp.message(AdminStates.waiting_for_setting_val)
async def st_setting_val(message: types.Message, state: FSMContext):
    val = safe_int(message.text, -1)
    if val < 0:
        await message.answer("Введите корректное число.")
        return
    
    data = await state.get_data()
    shift_num = data.get("shift_num")
    key = f"total_{shift_num}"
    
    set_setting(key, val)
    await state.clear()
    await message.answer(f"✅ Установлено {val} для {shift_num} смены.")
    await admin_render_main(message)

# --- Список подтвержденных (Улучшенный UI) ---
@dp.callback_query(F.data == "admin_approved")
async def cb_admin_approved(cb: types.CallbackQuery):
    if not is_admin(cb.from_user.id): return
    
    teachers = get_teachers_by_status(True)
    if not teachers:
        await cb.message.edit_text("✅ Список пуст.", reply_markup=back_kb())
        return

    # Формируем список кнопок (по 1 в ряд)
    kb_rows = []
    for t in teachers:
        lbl = format_teacher_label(t)
        kb_rows.append([InlineKeyboardButton(text=lbl, callback_data=f"manage_teacher|{t['id']}")])
    
    kb_rows.append([InlineKeyboardButton(text="🔙 Назад", callback_data="admin_back")])
    
    await cb.message.edit_text(
        "✅ <b>Подтвержденные учителя</b>\nНажмите на учителя для управления:", 
        reply_markup=InlineKeyboardMarkup(inline_keyboard=kb_rows)
    )

@dp.callback_query(F.data.startswith("manage_teacher|"))
async def cb_manage_teacher(cb: types.CallbackQuery):
    tid = int(cb.data.split("|")[1])
    t = get_teacher_by_id(tid)
    if not t:
        await safe_answer(cb, "Не найден", True)
        return
    
    uname = f"@{t['username']}" if t['username'] else "нет"
    text = (
        f"👤 <b>Карточка учителя</b>\n"
        f"ID: <code>{t['id']}</code>\n"
        f"ФИО: <b>{t['full_name']}</b>\n"
        f"User: {uname}\n"
        f"Класс: <b>{t.get('class_name') or '—'}</b>\n"
        f"Предмет: <b>{t.get('subject') or '—'}</b>"
    )
    await cb.message.edit_text(text, reply_markup=teacher_manage_kb(tid))

# --- Список ожидающих ---
@dp.callback_query(F.data == "admin_pending")
async def cb_admin_pending(cb: types.CallbackQuery):
    if not is_admin(cb.from_user.id): return
    
    teachers = get_teachers_by_status(False)
    if not teachers:
        await cb.message.edit_text("⏳ Заявок нет.", reply_markup=back_kb())
        return

    # Тоже выводим списком, но при клике сразу меню действий
    kb_rows = []
    for t in teachers:
        lbl = f"⏳ {t['full_name']} ({t.get('username') or '?'})"
        kb_rows.append([InlineKeyboardButton(text=lbl, callback_data=f"pending_teacher|{t['id']}")])
    
    kb_rows.append([InlineKeyboardButton(text="🔙 Назад", callback_data="admin_back")])
    await cb.message.edit_text("⏳ <b>Ожидают подтверждения</b>:", reply_markup=InlineKeyboardMarkup(inline_keyboard=kb_rows))

@dp.callback_query(F.data.startswith("pending_teacher|"))
async def cb_pending_details(cb: types.CallbackQuery):
    tid = int(cb.data.split("|")[1])
    t = get_teacher_by_id(tid)
    if not t: return
    
    uname = f"@{t['username']}" if t['username'] else "нет"
    text = (
        f"🆕 <b>Новая заявка</b>\n"
        f"ФИО: {t['full_name']}\n"
        f"User: {uname}\n"
        f"Telegram ID: <code>{t['telegram_id']}</code>"
    )
    await cb.message.edit_text(text, reply_markup=teacher_pending_kb(tid))

# --- Действия над учителем ---
@dp.callback_query(F.data.startswith(("appr|", "decl|")))
async def cb_teacher_action(cb: types.CallbackQuery):
    action, tid = cb.data.split("|")
    tid = int(tid)
    
    if action == "appr":
        approve_teacher_db(tid)
        await safe_answer(cb, "✅ Учитель подтвержден")
    else:
        decline_teacher_db(tid)
        await safe_answer(cb, "🗑 Учитель удален")
    
    # Возврат к списку
    await cb_admin_approved(cb) if action == "decl" else await cb_admin_pending(cb)

@dp.callback_query(F.data.startswith(("setclass|", "setsubj|")))
async def cb_edit_field(cb: types.CallbackQuery, state: FSMContext):
    action, tid = cb.data.split("|")
    field = "class_name" if action == "setclass" else "subject"
    
    await state.update_data(tid=tid, field=field)
    if field == "class_name":
        await state.set_state(AdminStates.waiting_for_class)
        await cb.message.answer("Введите класс (например 7А):")
    else:
        await state.set_state(AdminStates.waiting_for_subject)
        await cb.message.answer("Введите предмет:")
    await safe_answer(cb)

@dp.message(AdminStates.waiting_for_class)
@dp.message(AdminStates.waiting_for_subject)
async def st_save_field(message: types.Message, state: FSMContext):
    data = await state.get_data()
    tid = data.get("tid")
    field = data.get("field")
    val = (message.text or "").strip()
    
    update_teacher_field(tid, field, val)
    await state.clear()
    await message.answer(f"✅ Сохранено: {val}")
    
    # Возвращаемся в карточку
    t = get_teacher_by_id(tid)
    if t and is_admin(message.from_user.id):
        # Эмуляция возврата в админку (можно просто отправить текст)
        await admin_render_main(message)

# --- Поиск ---
@dp.callback_query(F.data == "admin_search")
async def cb_search_start(cb: types.CallbackQuery, state: FSMContext):
    await state.set_state(AdminStates.waiting_for_search)
    await cb.message.answer("Введите ФИО, класс или ID для поиска:")
    await safe_answer(cb)

@dp.message(AdminStates.waiting_for_search)
async def st_search_res(message: types.Message, state: FSMContext):
    q = message.text
    rows = search_teachers(q)
    await state.clear()
    
    if not rows:
        await message.answer("Ничего не найдено.")
        return

    kb_rows = []
    for t in rows:
        icon = "✅" if t['is_approved'] else "⏳"
        btn_txt = f"{icon} {t['full_name']} ({t.get('class_name') or '-'})"
        action = "manage_teacher" if t['is_approved'] else "pending_teacher"
        kb_rows.append([InlineKeyboardButton(text=btn_txt, callback_data=f"{action}|{t['id']}")])
    
    kb_rows.append([InlineKeyboardButton(text="🔙 Назад", callback_data="admin_back")])
    await message.answer(f"🔍 Результаты поиска ({len(rows)}):", reply_markup=InlineKeyboardMarkup(inline_keyboard=kb_rows))

# --- Ручная рассылка ---
@dp.callback_query(F.data == "admin_broadcast")
async def cb_broadcast(cb: types.CallbackQuery, state: FSMContext):
    await state.set_state(AdminStates.waiting_for_broadcast)
    await cb.message.answer("Введите текст сообщения для ВСЕХ учителей:")
    await safe_answer(cb)

@dp.message(AdminStates.waiting_for_broadcast)
async def st_broadcast_send(message: types.Message, state: FSMContext):
    txt = message.text
    await state.clear()
    
    teachers = get_teachers_by_status(True)
    cnt = 0
    for t in teachers:
        try:
            await bot.send_message(t['telegram_id'], f"📢 <b>Сообщение от администратора:</b>\n\n{txt}")
            cnt += 1
            await asyncio.sleep(0.05)
        except: pass
    
    await message.answer(f"✅ Рассылка отправлена {cnt} пользователям.")
    await admin_render_main(message)


# =========================
# USER COMMANDS
# =========================
@dp.message(Command("start"))
async def cmd_start(message: types.Message):
    # Обновляем юзернейм при старте
    update_username_if_changed(message.from_user.id, message.from_user.username)

    t = get_teacher_by_telegram(message.from_user.id)
    if t:
        if t['is_approved']:
            await message.answer("👋 С возвращением! Используйте /report для отчета.")
        else:
            await message.answer("⏳ Ваша заявка все еще на рассмотрении.")
    elif is_admin(message.from_user.id):
        await message.answer("👑 Вы админ. Жмите /admin")
    else:
        await message.answer("🔒 Введите секретный код для регистрации.")

@dp.message(Command("report"))
async def cmd_report(message: types.Message):
    await show_report_user(message)

async def show_report_user(message: types.Message):
    # Обновляем юзернейм при любом взаимодействии с отчетом
    update_username_if_changed(message.from_user.id, message.from_user.username)

    t = get_teacher_by_telegram(message.from_user.id)
    if not t or not t['is_approved']:
        await message.answer("⛔ Нет доступа.")
        return
    
    d = today_iso()
    rep = create_or_get_report(t['id'], d)
    
    if rep['is_submitted']:
        await message.answer("✅ Отчет на сегодня уже отправлен.")
    else:
        txt = (
            f"📅 Отчет за <b>{d}</b>\n"
            f"🤒 Болеют: {rep['sick']}\n"
            f"🏥 Санаторий: {rep['sanatorium']}\n"
            f"📄 Заявление: {rep['parent_statement']}\n"
            f"🏆 Соревнования: {rep['competition']}\n"
            f"❌ Без причины: {rep['absent_without_reason']}"
        )
        await message.answer(txt, reply_markup=report_kb(rep))

@dp.callback_query(F.data == "manual_report_start")
async def cb_manual_start(cb: types.CallbackQuery):
    await safe_answer(cb)
    await show_report_user(cb.message)

# Обработка кнопок отчета (edit, set, submit...)
@dp.callback_query(F.data.startswith(("edit|", "set|", "delta|", "back|", "refresh|", "submit|", "noop")))
async def cb_report_logic(cb: types.CallbackQuery):
    # Анти-спам
    if cb.data == "noop":
        await safe_answer(cb)
        return

    # Блокировка
    lock = USER_LOCKS[cb.from_user.id]
    if lock.locked():
        await safe_answer(cb, "⏳")
        return

    async with lock:
        parts = cb.data.split("|")
        action = parts[0]
        
        if action == "submit":
            rid = int(parts[1])
            if submit_report(rid):
                await safe_edit_text(cb.message, "✅ Отчет успешно отправлен!")
            else:
                await safe_answer(cb, "Ошибка или уже отправлено", True)
            return
        
        if action == "back" or action == "refresh":
            rid = int(parts[1])
            conn = get_connection()
            r = conn.execute("SELECT * FROM daily_reports WHERE id=?", (rid,)).fetchone()
            conn.close()
            if r:
                r_dict = dict(r)
                txt = (
                    f"📅 Отчет за <b>{r_dict['report_date']}</b>\n"
                    f"🤒 Болеют: {r_dict['sick']}\n"
                    f"🏥 Санаторий: {r_dict['sanatorium']}\n"
                    f"📄 Заявление: {r_dict['parent_statement']}\n"
                    f"🏆 Соревнования: {r_dict['competition']}\n"
                    f"❌ Без причины: {r_dict['absent_without_reason']}"
                )
                await safe_edit_text(cb.message, txt, report_kb(r_dict))
            return
            
        # Логика изменения чисел
        if action == "edit":
            field, rid = parts[1], int(parts[2])
            conn = get_connection()
            r = conn.execute("SELECT * FROM daily_reports WHERE id=?", (rid,)).fetchone()
            conn.close()
            if r and not r['is_submitted']:
                await safe_edit_text(
                    cb.message, 
                    f"Изменение поля: <b>{field}</b>\nТекущее: {r[field]}", 
                    number_kb(field, r[field], rid)
                )
            else:
                await safe_answer(cb, "Уже отправлено", True)
            return
        
        if action in ("set", "delta"):
            field = parts[1]
            rid = int(parts[3])
            
            conn = get_connection()
            r = conn.execute("SELECT * FROM daily_reports WHERE id=?", (rid,)).fetchone()
            conn.close()
            
            if not r or r['is_submitted']:
                await safe_answer(cb, "Недоступно", True)
                return

            current = r[field]
            val = int(parts[2])
            new_val = val if action == "set" else current + val
            
            set_report_value(rid, field, new_val)
            
            # Обновляем отображение клавиатуры
            await safe_edit_text(
                cb.message, 
                f"Изменение поля: <b>{field}</b>\nНовое значение: {max(0, new_val)}", 
                number_kb(field, max(0, new_val), rid)
            )

# =========================
# REGISTRATION
# =========================
@dp.message()
async def catch_secret_code(message: types.Message):
    if message.text == SECRET_CODE:
        uname = message.from_user.username
        t = create_teacher_if_missing(message.from_user.id, message.from_user.full_name, uname)
        if not t['is_approved']:
            await message.answer("✅ Заявка принята. Ожидайте подтверждения.")
            # Уведомляем админов
            for adm in ADMIN_IDS:
                try:
                    await bot.send_message(adm, f"🆕 Новый учитель: {t['full_name']} (@{uname})\n/admin -> Ожидают")
                except: pass
        else:
            await message.answer("✅ Вы уже подтверждены.")

# =========================
# MAIN
# =========================
async def main():
    init_db()
    logger.info("Bot started")
    asyncio.create_task(schedule_loop())
    await dp.start_polling(bot)

if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        pass
